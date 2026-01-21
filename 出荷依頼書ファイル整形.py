import sys
import tkinter as tk
from tkinter import messagebox
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image
import qrcode
import os
import tempfile
import shutil
from pathlib import Path
from datetime import datetime
import gc  # ★ 追加
import time  # ★ 追加

# 印刷ダイアログを表示するためのライブラリ
try:
    import win32com.client
    from win32com.client import constants
    WIN32_AVAILABLE = True
except ImportError:
    WIN32_AVAILABLE = False

from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D

def get_username():
    return os.getlogin()

def find_csv_path():
    """
    入荷CSVファイルのパスを候補パスから探索して返す。
    見つからない場合は None を返す。
    """
    username = get_username()
    possible_paths = [
        f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\入荷実績\\【標準】_入荷.csv",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\入荷実績\\【標準】_入荷.csv",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\入荷実績\\【標準】_入荷.csv"
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            return path
    return None

def create_qr_code(data, size=100):
    """
    渡された文字列dataからQRコードを生成し、一時ディレクトリにPNGで保存してパスを返す。
    """
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)
    qr_image = qr.make_image(fill_color="black", back_color="white")
    
    temp_dir = tempfile.gettempdir()
    temp_path = os.path.join(temp_dir, f"qr_{hash(data)}.png")
    qr_image.save(temp_path)
    return temp_path

def get_worker_name(code):
    """
    コードから担当者名を取得する。定義がなければコードのまま返す。
    """
    worker_dict = {
        "11": "細田　宗之介",
        "12": "平松　良太",
        "13": "坂上　敦士",
        "15": "山岡　正延",
        "16": "土田　周平",
        "TD": "東京支店対応"
    }
    return worker_dict.get(str(code), code)

def get_line_number(value):
    """
    ソートに利用する。文字列から数値へ変換できなければ float('inf') を返して並びの最後に。
    """
    try:
        return int(str(value))
    except (ValueError, TypeError):
        return float('inf')

# ピクセル(px)をポイント(pt)に変換する関数
# 1インチ=96px, 1インチ=72pt → 1px ≈ 0.75pt
def px_to_points(px):
    return px * 72.0 / 96.0

# ポイント(pt)からEMUへ変換する関数
def points_to_EMU(points):
    """
    1pt = 1/72inch, 1inch = 914400 EMU → 1pt = 12700 EMU
    """
    return int(points * 12700)

def add_image_center(ws, img, cell_address, center_x=True, center_y=True):
    """
    指定したセルに画像を貼り付ける。
    center_x=True の場合はセルの横方向中央に、
    center_y=True の場合はセルの縦方向中央に配置。
    """
    row = int(''.join([i for i in cell_address if i.isdigit()]))
    col_letter = ''.join([i for i in cell_address if i.isalpha()])
    col_num = openpyxl.utils.column_index_from_string(col_letter)

    # セルの列幅（Excelのカスタマイズ値）と行高さを取得
    col_width = ws.column_dimensions[col_letter].width
    if col_width is None:
        col_width = 8  # デフォルト相当

    # おおよそ 1 列幅 ≈ 7pt で計算
    col_width_points = col_width * 7

    row_height = ws.row_dimensions[row].height
    if row_height is None:
        row_height = 15  # デフォルト相当
    row_height_points = row_height

    # 画像の幅・高さ (pt) → EMU 変換
    img_width_emu = points_to_EMU(img.width)
    img_height_emu = points_to_EMU(img.height)

    # セルの幅・高さ (pt) → EMU 変換
    cell_width_emu = points_to_EMU(col_width_points)
    cell_height_emu = points_to_EMU(row_height_points)

    # 中央寄せオフセット計算
    if center_x:
        col_off = (cell_width_emu - img_width_emu) // 2
    else:
        col_off = 0

    if center_y:
        row_off = (cell_height_emu - img_height_emu) // 2
    else:
        row_off = 0

    # AnchorMarker 作成
    marker_from = AnchorMarker(
        col=col_num - 1,
        colOff=col_off,
        row=row - 1,
        rowOff=row_off
    )
    marker_to = AnchorMarker(
        col=col_num - 1,
        colOff=col_off + img_width_emu,
        row=row - 1,
        rowOff=row_off + img_height_emu
    )

    size = XDRPositiveSize2D(img_width_emu, img_height_emu)
    anchor = OneCellAnchor(_from=marker_from, ext=size)
    anchor._to = marker_to

    ws.add_image(img, anchor)

def get_save_directory(original_code):
    """
    H13セルの置換前の値(例: "11", "12", "13", "16")から、
    該当のディレクトリ候補をリストで返し、存在するパスを返す。
    該当パスがなければ None を返す。
    """
    username = get_username()

    candidates_map = {
        "11": [
            f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\出荷依頼書\\11_細田宗之介",
            f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\出荷依頼書\\11_細田宗之介",
            f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\出荷依頼書\\11_細田宗之介"
        ],
        "12": [
            f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\出荷依頼書\\12_平松良太",
            f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\出荷依頼書\\12_平松良太",
            f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\出荷依頼書\\12_平松良太"
        ],
        "13": [
            f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\出荷依頼書\\13_坂上敦士",
            f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\出荷依頼書\\13_坂上敦士",
            f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\出荷依頼書\\13_坂上敦士"
        ],
        "16": [
            f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\出荷依頼書\\16_土田周平",
            f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\出荷依頼書\\16_土田周平",
            f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\出荷依頼書\\16_土田周平"
        ],
    }

    if str(original_code) not in candidates_map:
        return None

    for path in candidates_map[str(original_code)]:
        if os.path.exists(path):
            return path
    return None

def process_file(input_file):
    """
    与えられたExcelファイルパス(input_file)を処理する関数。
    """
    if not os.path.isfile(input_file):
        messagebox.showerror("エラー", f"ファイルが存在しません: {input_file}")
        return

    temp_output_file = os.path.splitext(input_file)[0] + "_temp.xlsx"
    qr_images = []
    wb = None

    # フラグ: 正常終了したかどうか
    success_flag = False

    try:
        shutil.copy2(input_file, temp_output_file)
        wb = openpyxl.load_workbook(temp_output_file)
        ws = wb.active

        # 1) H13セル(置換前コード)
        original_code_cell = ws.cell(row=13, column=8)  # H13
        original_code_value = str(original_code_cell.value) if original_code_cell.value else ""

        # 2) 出荷依頼番号一覧から最小/最大取得
        shipping_numbers = []
        for row_obj in ws['G']:
            if row_obj.value == "出荷依頼番号：":
                num = ws.cell(row=row_obj.row, column=row_obj.column + 1).value
                if num is not None:
                    shipping_numbers.append(str(num))

        if len(shipping_numbers) == 0:
            messagebox.showerror("エラー", "シート上に出荷依頼番号が見つかりません。")
            return

        min_ship_num = min(shipping_numbers)
        max_ship_num = max(shipping_numbers)
        today_str = datetime.now().strftime('%y%m%d')
        output_filename = f"{min_ship_num}_{max_ship_num}_{today_str}.xlsx"

        # 3) 作業担当者置換
        for row_obj in ws['G']:
            if row_obj.value == "作業担当者：":
                code_cell = ws.cell(row=row_obj.row, column=row_obj.column + 1)
                code_cell.value = get_worker_name(str(code_cell.value))

        # 4) 行番号_L のソート
        line_number_rows = []
        for row_obj in ws['B']:
            if row_obj.value == "行番号_L":
                line_number_rows.append(row_obj.row)

        if not line_number_rows:
            messagebox.showerror("エラー", "行番号_Lが見つかりません")
            return

        for line_number_row in line_number_rows:
            sort_data = []
            current_row = line_number_row + 1
            while ws.cell(row=current_row, column=2).value is not None:
                row_data = []
                for col in range(1, ws.max_column + 1):
                    row_data.append(ws.cell(row=current_row, column=col).value)
                sort_data.append(row_data)
                current_row += 1

            n = len(sort_data)
            for i in range(n):
                for j in range(n - i - 1):
                    current_value = get_line_number(sort_data[j][1])
                    next_value = get_line_number(sort_data[j + 1][1])
                    if current_value > next_value:
                        sort_data[j], sort_data[j + 1] = sort_data[j + 1], sort_data[j]

            # ソート結果を反映
            for i, row_data in enumerate(sort_data):
                for j, value in enumerate(row_data):
                    ws.cell(row=line_number_row + 1 + i, column=j + 1).value = value

        # 5) 入荷CSV読み込み
        csv_path = find_csv_path()
        if not csv_path:
            messagebox.showerror("エラー", "入荷CSVファイルが見つかりません")
            return

        df_nyuka = pd.read_csv(csv_path, encoding='cp932', low_memory=False)

        # 発注番号の処理
        order_number_rows = []
        for row_obj in ws['G']:
            if row_obj.value == "発注番号":
                order_number_rows.append(row_obj.row)

        for order_number_row in order_number_rows:
            current_row = order_number_row + 1
            while ws.cell(row=current_row, column=6).value is not None:
                lot_number = str(ws.cell(row=current_row, column=6).value)
                matching_row = df_nyuka[df_nyuka['明細_ロット番号'] == lot_number]
                if not matching_row.empty:
                    ws.cell(row=current_row, column=7).value = matching_row.iloc[0]['発注番号']
                current_row += 1

        # 6) ロット番号QR (ロット番号 + '_' + 出荷数量) 85px（行高さギリギリ）
        lot_number_rows = []
        for row_obj in ws['F']:
            if row_obj.value == "ロット番号":
                lot_number_rows.append(row_obj.row)

        for lot_number_row in lot_number_rows:
            current_row = lot_number_row + 1
            while ws.cell(row=current_row, column=6).value is not None:
                # ★ 明細行の高さを66.00（88ピクセル）に設定
                ws.row_dimensions[current_row].height = 66.00
                
                lot_number = str(ws.cell(row=current_row, column=6).value)
                shipping_qty = ws.cell(row=current_row, column=5).value
                
                # 変更ポイント: shipping_qtyを3桁ゼロパディング
                if shipping_qty is None:
                    # 数量が未入力の場合は "000" とするか、もしくは空文字にするなど運用に応じて
                    shipping_qty_str = "000"
                else:
                    # 数値変換に失敗したら例外処理を入れる
                    try:
                        shipping_qty_str = str(int(shipping_qty)).zfill(3)  # 3桁ゼロパディング
                    except ValueError:
                        # 変換できない場合は別の文字を入れるなど対応
                        shipping_qty_str = "000"
                
                # QRコードに埋め込む文字列
                qr_data = f"{lot_number}_{shipping_qty_str}"
                
                qr_path = create_qr_code(qr_data)
                qr_images.append(qr_path)
                img = Image(qr_path)

                # ★ QRコードサイズを85px（行高さギリギリ）に変更
                px85 = px_to_points(85)
                img.width = px85
                img.height = px85

                col_letter = 'A' if (current_row - lot_number_row) % 2 == 1 else 'H'
                cell = f'{col_letter}{current_row}'
                add_image_center(ws, img, cell, center_x=True, center_y=True)
                current_row += 1

        # 7) 出荷依頼番号QR (横のみ中央) 85pxに拡大
        for row_obj in ws['G']:
            if row_obj.value == "出荷依頼番号：":
                shipping_number = str(ws.cell(row=row_obj.row, column=row_obj.column + 1).value)
                qr_path = create_qr_code(shipping_number)
                qr_images.append(qr_path)
                img = Image(qr_path)

                # ★ QRコードサイズを85pxに変更（ロット番号QRと統一）
                px85 = px_to_points(85)
                img.width = px85
                img.height = px85

                cell = f'H{row_obj.row + 1}'
                add_image_center(ws, img, cell, center_x=True, center_y=False)

        # 8) 保存先ディレクトリの特定
        save_dir = get_save_directory(original_code_value)
        if not save_dir:
            messagebox.showerror("エラー", f"H13セルの値({original_code_value})に対応する保存先フォルダが見つかりません。")
            return

        # 8-1) "YYMMDD" フォルダを作成 (なければ)
        date_folder = os.path.join(save_dir, today_str)
        if not os.path.exists(date_folder):
            os.makedirs(date_folder)

        final_output_file = os.path.join(date_folder, output_filename)

        # 9) 保存
        wb.save(final_output_file)
        wb.close()
        wb = None

        # ★ ガベージコレクション強制実行とファイル解放待機
        gc.collect()  # メモリ解放
        time.sleep(1)  # 1秒待機

        if os.path.exists(temp_output_file):
            os.remove(temp_output_file)

        # 成功フラグを立てる
        success_flag = True

        # ★ 元ファイル削除もガベージコレクション後に実行
        if os.path.exists(input_file):
            try:
                gc.collect()  # 再度ガベージコレクション
                time.sleep(0.5)  # 少し待機
                os.remove(input_file)
            except Exception as e:
                messagebox.showerror("ファイル削除エラー", f"元ファイルを削除できませんでした: {input_file}\n{str(e)}")

        # 10) Excelを開き 印刷ダイアログを表示
        if WIN32_AVAILABLE:
            try:
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = True
                wb2 = excel.Workbooks.Open(final_output_file)
                
                # メッセージボックスで確認を表示
                result = messagebox.askyesno("確認", "ファイルが開かれました。印刷プレビューを表示しますか？")
                if result:
                    # ユーザーが確認した場合のみ印刷プレビューを試みる
                    wb2.Activate()
                    excel.ActiveSheet.PrintPreview()
            except Exception as ee:
                messagebox.showerror("エラー", f"操作中にエラーが発生しました:\n{ee}")
        else:
            try:
                os.startfile(final_output_file)
                messagebox.showinfo("情報", "pywin32 がインストールされていないため、印刷ダイアログは表示できません。\nファイルを開きましたので、手動で印刷してください。")
            except:
                pass

        messagebox.showinfo(
            "完了",
            f"処理が完了しました。\n保存先:\n{final_output_file}"
        )

    except Exception as e:
        messagebox.showerror("エラー", str(e))
        if os.path.exists(temp_output_file):
            try:
                os.remove(temp_output_file)
            except:
                pass
    finally:
        if wb:
            try:
                wb.close()
            except:
                pass

        for qr_path in qr_images:
            try:
                if os.path.exists(qr_path):
                    os.remove(qr_path)
            except:
                pass

        # ★ finally句からは元ファイル削除処理を削除
        # （上記で既に削除済みのため）

        # ★ 処理が成功した場合のみ、元ファイルを削除
        if success_flag:
            if os.path.exists(input_file):
                try:
                    os.remove(input_file)
                except Exception as e:
                    messagebox.showerror("ファイル削除エラー", f"元ファイルを削除できませんでした: {input_file}\n{str(e)}")

def main():
    """
    メイン処理：スクリプトにドラッグ＆ドロップされたファイルを処理。
    """
    root = tk.Tk()
    root.withdraw()

    if len(sys.argv) < 2:
        messagebox.showinfo("情報", "このファイルにExcelをドラッグ＆ドロップするか、『送る』に設定して実行してください。")
        return

    for file_path in sys.argv[1:]:
        process_file(file_path)

if __name__ == "__main__":
    main()
